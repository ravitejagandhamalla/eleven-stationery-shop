import os
import psycopg2
from openpyxl import Workbook
from flask import send_file
from reportlab.platypus import SimpleDocTemplate, Paragraph, Spacer, Table
from reportlab.lib import colors
from reportlab.lib.styles import getSampleStyleSheet
from reportlab.lib import pagesizes
import io
import io
from flask import Flask, render_template, request, redirect, url_for, session, flash

app = Flask(__name__)
app.secret_key = "supersecretkey"


# ==============================
# DATABASE CONNECTION
# ==============================
def get_db_connection():
    database_url = os.environ.get("DATABASE_URL")
    if not database_url:
        raise RuntimeError("DATABASE_URL is not set")

    return psycopg2.connect(database_url)


# ==============================
# HOME
# ==============================
@app.route("/")
def index():
    if "user_id" in session:
        return redirect(url_for("dashboard"))
    return redirect(url_for("login"))


# ==============================
# LOGIN
# ==============================
@app.route("/login", methods=["GET", "POST"])
def login():
    if request.method == "POST":
        username = request.form["username"]
        password = request.form["password"]

        conn = get_db_connection()
        cur = conn.cursor()

        cur.execute(
            "SELECT id FROM users WHERE username=%s AND password=%s",
            (username, password),
        )
        user = cur.fetchone()

        cur.close()
        conn.close()

        if user:
            session["user_id"] = user[0]
            session["username"] = username
            return redirect(url_for("dashboard"))
        else:
            flash("Invalid username or password")

    return render_template("login.html")


# ==============================
# LOGOUT
# ==============================
@app.route("/logout")
def logout():
    session.clear()
    return redirect(url_for("login"))


# ==============================
# DASHBOARD
# ==============================
@app.route("/dashboard")
def dashboard():
    if "user_id" not in session:
        return redirect(url_for("login"))

    conn = get_db_connection()
    cur = conn.cursor()

    cur.execute("""
        SELECT COALESCE(SUM(amount),0)
        FROM income
        WHERE user_id=%s AND date = CURRENT_DATE
    """, (session["user_id"],))
    today_income = cur.fetchone()[0]

    cur.execute("""
        SELECT COALESCE(SUM(amount),0)
        FROM expenses
        WHERE user_id=%s AND date = CURRENT_DATE
    """, (session["user_id"],))
    today_expense = cur.fetchone()[0]

    cur.execute("""
        SELECT COUNT(*)
        FROM income
        WHERE user_id=%s
    """, (session["user_id"],))
    total_items = cur.fetchone()[0]

    cur.execute("""
        SELECT COALESCE(SUM(amount),0)
        FROM income
        WHERE user_id=%s
        AND date_trunc('month', date) = date_trunc('month', CURRENT_DATE)
    """, (session["user_id"],))
    monthly_income = cur.fetchone()[0]

    cur.execute("""
        SELECT COALESCE(SUM(amount),0)
        FROM expenses
        WHERE user_id=%s
        AND date_trunc('month', date) = date_trunc('month', CURRENT_DATE)
    """, (session["user_id"],))
    monthly_expense = cur.fetchone()[0]

    monthly_profit = monthly_income - monthly_expense

    cur.close()
    conn.close()

    return render_template(
        "dashboard.html",
        today_income=today_income,
        today_expense=today_expense,
        total_items=total_items,
        monthly_profit=monthly_profit,
    )


# ==============================
# ADD INCOME
# ==============================
@app.route("/income", methods=["GET", "POST"])
def income():
    if "user_id" not in session:
        return redirect(url_for("login"))

    if request.method == "POST":
        amount = request.form["amount"]
        description = request.form["description"]

        conn = get_db_connection()
        cur = conn.cursor()

        cur.execute(
            "INSERT INTO income (user_id, amount, description) VALUES (%s,%s,%s)",
            (session["user_id"], amount, description),
        )

        conn.commit()
        cur.close()
        conn.close()

        flash("Income added successfully")
        return redirect(url_for("dashboard"))

    return render_template("income.html")


# ==============================
# ADD EXPENSE
# ==============================
@app.route("/expenses", methods=["GET", "POST"])
def expenses():
    if "user_id" not in session:
        return redirect(url_for("login"))

    conn = get_db_connection()
    cur = conn.cursor()

    if request.method == "POST":
        date = request.form["date"]
        amount = request.form["amount"]
        purpose = request.form["purpose"]

        cur.execute("""
            INSERT INTO expenses (user_id, date, amount, purpose)
            VALUES (%s, %s, %s, %s)
        """, (session["user_id"], date, amount, purpose))

        conn.commit()
        cur.close()
        conn.close()

        return redirect(url_for("dashboard"))

    cur.close()
    conn.close()
    return render_template("expenses.html")


# ==============================
# VIEW RECORDS
# ==============================
@app.route("/view_records")
def view_records():
    if "user_id" not in session:
        return redirect(url_for("login"))

    record_type = request.args.get("type", "both")
    start_date = request.args.get("start")
    end_date = request.args.get("end")

    conn = get_db_connection()
    cur = conn.cursor()

    incomes = []
    expenses = []

    # =========================
    # INCOME FILTER
    # =========================
    if record_type in ["income", "both"]:
        query = """
            SELECT id, date, amount, description
            FROM income
            WHERE user_id=%s
        """
        params = [session["user_id"]]

        if start_date and end_date:
            query += " AND date BETWEEN %s AND %s"
            params.extend([start_date, end_date])

        query += " ORDER BY date DESC"

        cur.execute(query, tuple(params))
        incomes = cur.fetchall()

    # =========================
    # EXPENSE FILTER
    # =========================
    if record_type in ["expenses", "both"]:
        query = """
            SELECT id, date, amount, purpose
            FROM expenses
            WHERE user_id=%s
        """
        params = [session["user_id"]]

        if start_date and end_date:
            query += " AND date BETWEEN %s AND %s"
            params.extend([start_date, end_date])

        query += " ORDER BY date DESC"

        cur.execute(query, tuple(params))
        expenses = cur.fetchall()

    cur.close()
    conn.close()

    return render_template(
        "view_records.html",
        incomes=incomes,
        expenses=expenses,
        record_type=record_type,
        start=start_date,
        end=end_date
    )


# ==============================
# EDIT INCOME
# ==============================
@app.route("/edit_income/<int:id>", methods=["POST"])
def edit_income(id):
    if "user_id" not in session:
        return redirect(url_for("login"))

    amount = request.form["amount"]
    description = request.form["description"]

    conn = get_db_connection()
    cur = conn.cursor()

    cur.execute(
        "UPDATE income SET amount=%s, description=%s WHERE id=%s AND user_id=%s",
        (amount, description, id, session["user_id"]),
    )

    conn.commit()
    cur.close()
    conn.close()

    return redirect(url_for("view_records"))


# ==============================
# DELETE INCOME
# ==============================
@app.route("/delete_income/<int:id>")
def delete_income(id):
    if "user_id" not in session:
        return redirect(url_for("login"))

    conn = get_db_connection()
    cur = conn.cursor()

    cur.execute(
        "DELETE FROM income WHERE id=%s AND user_id=%s",
        (id, session["user_id"]),
    )

    conn.commit()
    cur.close()
    conn.close()

    return redirect(url_for("view_records"))


# ==============================
# EDIT EXPENSE
# ==============================
@app.route("/edit_expenses/<int:id>", methods=["POST"])
def edit_expenses(id):
    if "user_id" not in session:
        return redirect(url_for("login"))

    amount = request.form["amount"]
    purpose = request.form["purpose"]

    conn = get_db_connection()
    cur = conn.cursor()

    cur.execute(
        "UPDATE expenses SET amount=%s, purpose=%s WHERE id=%s AND user_id=%s",
        (amount, purpose, id, session["user_id"]),
    )

    conn.commit()
    cur.close()
    conn.close()

    return redirect(url_for("view_records"))


# ==============================
# DELETE EXPENSE
# ==============================
@app.route("/delete_expense/<int:id>")
def delete_expense(id):
    if "user_id" not in session:
        return redirect(url_for("login"))

    conn = get_db_connection()
    cur = conn.cursor()

    cur.execute(
        "DELETE FROM expenses WHERE id=%s AND user_id=%s",
        (id, session["user_id"]),
    )

    conn.commit()
    cur.close()
    conn.close()

    return redirect(url_for("view_records"))


# ==============================
# SUMMARY
# ==============================
@app.route("/summary")
def summary():
    if "user_id" not in session:
        return redirect(url_for("login"))

    conn = get_db_connection()
    cur = conn.cursor()

    # Total Income
    cur.execute("""
        SELECT COALESCE(SUM(amount),0)
        FROM income
        WHERE user_id=%s
    """, (session["user_id"],))
    total_income = cur.fetchone()[0]

    # Total Expense
    cur.execute("""
        SELECT COALESCE(SUM(amount),0)
        FROM expenses
        WHERE user_id=%s
    """, (session["user_id"],))
    total_expense = cur.fetchone()[0]

    profit = total_income - total_expense

    # ==========================
    # DAILY BREAKDOWN QUERY
    # ==========================

    cur.execute("""
        SELECT d.date,
               COALESCE(i.total_income, 0),
               COALESCE(e.total_expense, 0),
               COALESCE(i.total_income, 0) - COALESCE(e.total_expense, 0) AS daily_profit
        FROM (
            SELECT date FROM income WHERE user_id=%s
            UNION
            SELECT date FROM expenses WHERE user_id=%s
        ) d
        LEFT JOIN (
            SELECT date, SUM(amount) AS total_income
            FROM income
            WHERE user_id=%s
            GROUP BY date
        ) i ON d.date = i.date
        LEFT JOIN (
            SELECT date, SUM(amount) AS total_expense
            FROM expenses
            WHERE user_id=%s
            GROUP BY date
        ) e ON d.date = e.date
        ORDER BY d.date DESC
    """, (
        session["user_id"],
        session["user_id"],
        session["user_id"],
        session["user_id"]
    ))

    daily_data = cur.fetchall()

    cur.close()
    conn.close()

    return render_template(
        "summary.html",
        total_income=total_income,
        total_expense=total_expense,
        profit=profit,
        daily_data=daily_data
    )


# ==============================
# CHANGE PASSWORD
# ==============================
@app.route("/change_password", methods=["GET", "POST"])
def change_password():
    if "user_id" not in session:
        return redirect(url_for("login"))

    if request.method == "POST":
        new_password = request.form["new_password"]

        conn = get_db_connection()
        cur = conn.cursor()

        cur.execute(
            "UPDATE users SET password=%s WHERE id=%s",
            (new_password, session["user_id"]),
        )

        conn.commit()
        cur.close()
        conn.close()

        flash("Password changed successfully")
        return redirect(url_for("dashboard"))

    return render_template("change_password.html")


# ==============================
# FORGOT PASSWORD
# ==============================
@app.route("/forgot_password", methods=["GET", "POST"])
def forgot_password():
    if request.method == "POST":
        username = request.form.get("username")
        new_password = request.form.get("new_password")

        if not username or not new_password:
            flash("All fields required")
            return redirect(url_for("forgot_password"))

        conn = get_db_connection()
        cur = conn.cursor()

        cur.execute(
            "UPDATE users SET password=%s WHERE username=%s",
            (new_password, username),
        )

        conn.commit()
        cur.close()
        conn.close()

        flash("Password updated successfully")
        return redirect(url_for("login"))

    return render_template("forgot_password.html")
    # ==============================
# EXPORT TO EXCEL
# ==============================
@app.route("/export_excel")
def export_excel():
    if "user_id" not in session:
        return redirect(url_for("login"))

    conn = get_db_connection()
    cur = conn.cursor()

    # Get income records
    cur.execute("""
        SELECT date, amount, description
        FROM income
        WHERE user_id=%s
        ORDER BY date DESC
    """, (session["user_id"],))
    incomes = cur.fetchall()

    # Get expense records
    cur.execute("""
        SELECT date, amount, purpose
        FROM expenses
        WHERE user_id=%s
        ORDER BY date DESC
    """, (session["user_id"],))
    expenses = cur.fetchall()

    cur.close()
    conn.close()

    wb = Workbook()
    ws = wb.active
    ws.title = "Records"

    ws.append(["Type", "Date", "Amount", "Description"])

    for i in incomes:
        ws.append(["Income", i[0], i[1], i[2]])

    for e in expenses:
        ws.append(["Expense", e[0], e[1], e[2]])

    file_stream = io.BytesIO()
    wb.save(file_stream)
    file_stream.seek(0)

    return send_file(
        file_stream,
        as_attachment=True,
        download_name="records.xlsx",
        mimetype="application/vnd.openxmlformats-officedocument.spreadsheetml.sheet"
    )
    # ==================================
# EXPORT FILTERED EXCEL (DATE RANGE)
# ==================================
@app.route("/export_filtered_excel")
def export_filtered_excel():
    if "user_id" not in session:
        return redirect(url_for("login"))

    start = request.args.get("start")
    end = request.args.get("end")

    conn = get_db_connection()
    cur = conn.cursor()

    # Income filter
    income_query = """
        SELECT date, amount, description
        FROM income
        WHERE user_id=%s
    """
    expense_query = """
        SELECT date, amount, purpose
        FROM expenses
        WHERE user_id=%s
    """

    params = [session["user_id"]]

    if start and end:
        income_query += " AND date BETWEEN %s AND %s"
        expense_query += " AND date BETWEEN %s AND %s"
        params += [start, end]

    income_query += " ORDER BY date DESC"
    expense_query += " ORDER BY date DESC"

    cur.execute(income_query, params)
    incomes = cur.fetchall()

    cur.execute(expense_query, params)
    expenses = cur.fetchall()

    cur.close()
    conn.close()

    wb = Workbook()

    # ---------------- INCOME SHEET ----------------
    ws1 = wb.active
    ws1.title = "Income"
    ws1.append(["Date", "Amount", "Description"])

    for row in incomes:
        ws1.append(row)

    # ---------------- EXPENSE SHEET ----------------
    ws2 = wb.create_sheet("Expenses")
    ws2.append(["Date", "Amount", "Purpose"])

    for row in expenses:
        ws2.append(row)

    file_stream = io.BytesIO()
    wb.save(file_stream)
    file_stream.seek(0)

    return send_file(
        file_stream,
        as_attachment=True,
        download_name="filtered_records.xlsx",
        mimetype="application/vnd.openxmlformats-officedocument.spreadsheetml.sheet"
    )
    # ==================================
# EXPORT TO PDF
# ==================================
@app.route("/export_pdf")
def export_pdf():
    if "user_id" not in session:
        return redirect(url_for("login"))

    conn = get_db_connection()
    cur = conn.cursor()

    cur.execute("SELECT date, amount FROM income WHERE user_id=%s", (session["user_id"],))
    incomes = cur.fetchall()

    cur.execute("SELECT date, amount FROM expenses WHERE user_id=%s", (session["user_id"],))
    expenses = cur.fetchall()

    cur.close()
    conn.close()

    buffer = io.BytesIO()
    doc = SimpleDocTemplate(buffer, pagesize=pagesizes.A4)
    elements = []

    styles = getSampleStyleSheet()
    elements.append(Paragraph("ELEVEN STATIONARY SHOP - Report", styles["Title"]))
    elements.append(Spacer(1, 20))

    data = [["Type", "Date", "Amount"]]

    for i in incomes:
        data.append(["Income", str(i[0]), str(i[1])])

    for e in expenses:
        data.append(["Expense", str(e[0]), str(e[1])])

    table = Table(data)
    table.setStyle([
        ('BACKGROUND', (0,0), (-1,0), colors.grey),
        ('GRID', (0,0), (-1,-1), 1, colors.black)
    ])

    elements.append(table)
    doc.build(elements)

    buffer.seek(0)

    return send_file(
        buffer,
        as_attachment=True,
        download_name="records.pdf",
        mimetype="application/pdf"
    )
    # ==================================
# DOWNLOAD SUMMARY REPORT
# ==================================
@app.route("/download_summary")
def download_summary():
    if "user_id" not in session:
        return redirect(url_for("login"))

    conn = get_db_connection()
    cur = conn.cursor()

    cur.execute("SELECT COALESCE(SUM(amount),0) FROM income WHERE user_id=%s",
                (session["user_id"],))
    total_income = cur.fetchone()[0]

    cur.execute("SELECT COALESCE(SUM(amount),0) FROM expenses WHERE user_id=%s",
                (session["user_id"],))
    total_expense = cur.fetchone()[0]

    profit = total_income - total_expense

    cur.close()
    conn.close()

    buffer = io.BytesIO()
    doc = SimpleDocTemplate(buffer, pagesize=pagesizes.A4)
    elements = []

    styles = getSampleStyleSheet()

    elements.append(Paragraph("SUMMARY REPORT", styles["Title"]))
    elements.append(Spacer(1, 20))
    elements.append(Paragraph(f"Total Income: ₹ {total_income}", styles["Normal"]))
    elements.append(Paragraph(f"Total Expense: ₹ {total_expense}", styles["Normal"]))
    elements.append(Paragraph(f"Profit: ₹ {profit}", styles["Normal"]))

    doc.build(elements)

    buffer.seek(0)

    return send_file(
        buffer,
        as_attachment=True,
        download_name="summary_report.pdf",
        mimetype="application/pdf"
    )



# ==============================
# MAIN
# ==============================
if __name__ == "__main__":
    app.run(debug=True)
